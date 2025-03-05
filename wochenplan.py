import os
import json
import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import pygame

# MP3-Dateipfad
uwu_path = os.path.join(os.getenv('USERPROFILE'), 'Desktop', 'cute-uwu.mp3')

def get_top_left_cell(sheet, cell):
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell


def fill_cell(sheet, cell, value):
    top_left_cell = get_top_left_cell(sheet, cell)
    top_left_cell.value = value


def add_hyperlink(sheet, cell, link):
    cell = get_top_left_cell(sheet, cell)
    sheet[cell.coordinate] = f'=HYPERLINK("{link}", "Link")'


def fill_weekly_plan(input_data, template_path, output_path):
    wb = load_workbook(template_path)
    sheet = wb['Tabelle1']

    fill_cell(sheet, sheet['B3'], input_data["name"])
    fill_cell(sheet, sheet['E3'], input_data["vorname"])
    fill_cell(sheet, sheet['B4'], input_data["lehrgang"])
    fill_cell(sheet, sheet['D5'], input_data["kw_von"])
    fill_cell(sheet, sheet['F5'], input_data["kw_bis"])
    fill_cell(sheet, sheet['B5'], input_data["kalenderwoche"])

    weekdays = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag"]

    # Zeile bei der das Ausfüllen beginnt
    row_start = 9

    for i, day in enumerate(weekdays):
        if day in input_data["wochenplan"]:
            for j, entry in enumerate(input_data["wochenplan"][day]):
                row_offset = row_start + i * 9 + j  # Berechne die Zeile dynamisch
                fill_cell(sheet, sheet.cell(row=row_offset, column=3), entry.get("inhalt", ""))

                if "link" in entry:
                    add_hyperlink(sheet, sheet.cell(row=row_offset, column=8), entry["link"])

    wb.save(output_path)
    print(f"Datei wurde gespeichert unter: {output_path}")


def fetch_headings_from_url(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    headings = soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6'])
    formatted_data = []
    seen_headings = set()  # Set für doppelte Werte

    excluded_headings = ["Inhaltsverzeichnis", "Table of Contents", "Contents", "Index"]

    for i, heading in enumerate(headings, start=1):
        heading_text = heading.text.strip()

        if any(excluded_word in heading_text for excluded_word in excluded_headings):
            continue

        if heading_text and heading_text not in seen_headings:
            formatted_data.append({
                "inhalt": heading_text,
                "link": url  # Link wird hinzugefügt
            })
            seen_headings.add(heading_text)

        # Stoppen bei 9 Inhalten
        if len(formatted_data) >= 9:
            break

    return formatted_data


def run_gui():
    def on_submit():
        name = entry_name.get()
        vorname = entry_vorname.get()
        kw_von = entry_kw_von.get()
        kw_bis = entry_kw_bis.get()
        kalenderwoche = entry_kalenderwoche.get()

        links = {
            "Montag": entry_link_montag.get(),
            "Dienstag": entry_link_dienstag.get(),
            "Mittwoch": entry_link_mittwoch.get(),
            "Donnerstag": entry_link_donnerstag.get(),
            "Freitag": entry_link_freitag.get()
        }

        headings_data_by_day = {}
        for day, link in links.items():
            print(f"Hole Daten für {day} von {link}")
            headings_data_by_day[day] = fetch_headings_from_url(link)

        # Speichern in JSON
        with open('headings.json', 'w', encoding='utf-8') as f:
            json.dump(headings_data_by_day, f, ensure_ascii=False, indent=4)

        data = {
            "name": name,
            "vorname": vorname,
            "lehrgang": "nb-i",
            "kw_von": kw_von,
            "kw_bis": kw_bis,
            "kalenderwoche": kalenderwoche,
            "wochenplan": {
                "Montag": [{"inhalt": entry["inhalt"], "link": links["Montag"]} for entry in headings_data_by_day["Montag"]],
                "Dienstag": [{"inhalt": entry["inhalt"], "link": links["Dienstag"]} for entry in headings_data_by_day["Dienstag"]],
                "Mittwoch": [{"inhalt": entry["inhalt"], "link": links["Mittwoch"]} for entry in headings_data_by_day["Mittwoch"]],
                "Donnerstag": [{"inhalt": entry["inhalt"], "link": links["Donnerstag"]} for entry in headings_data_by_day["Donnerstag"]],
                "Freitag": [{"inhalt": entry["inhalt"], "link": links["Freitag"]} for entry in headings_data_by_day["Freitag"]]
            }
        }

        # Pfade für Vorlage und ausgefüllt
        template_path = os.path.join(os.getenv('USERPROFILE'), 'Desktop', 'Vorlage_Wochennachweis_9h.xlsx')
        output_path = os.path.join(os.getenv('USERPROFILE'), 'Desktop', 'Wochennachweis_ausgefuellt.xlsx')

        fill_weekly_plan(data, template_path, output_path)

        pygame.mixer.init()
        pygame.mixer.music.load(uwu_path)
        pygame.mixer.music.play()

    root = ttk.Window(themename="darkly")
    root.title("Wochennachweis Ausfüllen")

    # Eingabe für Name, Vorname, Kalenderwoche, KW-Von und KW-Bis
    ttk.Label(root, text="Name:").grid(row=0, column=0, padx=10, pady=5)
    entry_name = ttk.Entry(root)
    entry_name.grid(row=0, column=1, padx=10, pady=5)

    ttk.Label(root, text="Vorname:").grid(row=1, column=0, padx=10, pady=5)
    entry_vorname = ttk.Entry(root)
    entry_vorname.grid(row=1, column=1, padx=10, pady=5)

    ttk.Label(root, text="Kalenderwoche:").grid(row=2, column=0, padx=10, pady=5)
    entry_kalenderwoche = ttk.Entry(root)
    entry_kalenderwoche.grid(row=2, column=1, padx=10, pady=5)

    ttk.Label(root, text="KW von:").grid(row=3, column=0, padx=10, pady=5)
    entry_kw_von = ttk.Entry(root)
    entry_kw_von.grid(row=3, column=1, padx=10, pady=5)

    ttk.Label(root, text="KW bis:").grid(row=4, column=0, padx=10, pady=5)
    entry_kw_bis = ttk.Entry(root)
    entry_kw_bis.grid(row=4, column=1, padx=10, pady=5)

    ttk.Label(root, text="Link für Montag:").grid(row=5, column=0, padx=10, pady=5)
    entry_link_montag = ttk.Entry(root)
    entry_link_montag.grid(row=5, column=1, padx=10, pady=5)

    ttk.Label(root, text="Link für Dienstag:").grid(row=6, column=0, padx=10, pady=5)
    entry_link_dienstag = ttk.Entry(root)
    entry_link_dienstag.grid(row=6, column=1, padx=10, pady=5)

    ttk.Label(root, text="Link für Mittwoch:").grid(row=7, column=0, padx=10, pady=5)
    entry_link_mittwoch = ttk.Entry(root)
    entry_link_mittwoch.grid(row=7, column=1, padx=10, pady=5)

    ttk.Label(root, text="Link für Donnerstag:").grid(row=8, column=0, padx=10, pady=5)
    entry_link_donnerstag = ttk.Entry(root)
    entry_link_donnerstag.grid(row=8, column=1, padx=10, pady=5)

    ttk.Label(root, text="Link für Freitag:").grid(row=9, column=0, padx=10, pady=5)
    entry_link_freitag = ttk.Entry(root)
    entry_link_freitag.grid(row=9, column=1, padx=10, pady=5)

    submit_button = ttk.Button(root, text="Wochennachweis erstellen", command=on_submit)
    submit_button.grid(row=10, columnspan=2, pady=20)

    root.mainloop()


if __name__ == "__main__":
    run_gui()
